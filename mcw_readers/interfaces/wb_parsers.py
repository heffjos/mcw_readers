import re

from pathlib import Path
from datetime import datetime
from collections import namedtuple, defaultdict

import openpyxl

import numpy as np
import pandas as pd

try:
    import importlib.resources as pkg_resources
except ImportError:
    import importlib_resources as pkg_resources

from mcw_readers.interfaces.lut import lut
from mcw_readers import data
from mcw_readers.utils import DF_PSYCHOMETRIC, DICT_PSYCHOMETRIC, close_any

line = namedtuple('line', 'identifier test_no row')

class PedsParserError(Exception):
    """Exception raised if there is an error while parsing a peds neuroscore file"""
    pass

def read_workbook(fname):
    if fname.name.endswith('.csv'):
        df = pd.read_csv(fname)
    elif fname.name.endswith('.tsv'):
        df = pd.read_csv(fname, sep='\t')
    elif fname.name.endswith('.xlsx'):
        df = pd.read_excel(fname)
    else:
        raise Exception(f'Cannot parser spreadsheet: {fname}')

    return df

def peds_determine_variable_value(value, rc_variables, percentile):
    """
    Determines the variable value for 'ss' column
    
    Parameters
    ----------
    
    value : float
        the cell value
    rc_variables : list of str
        the list of redcap variables for the row of cell
    percentile : float
        the percentile value for the current row.
    
    Returns
    -------
    
    rc_variable : str
        the redcap variable name
    postprocessed_value : str or float
        the postprocessed cell value
    
    Description
    -----------
    
    Here are the lut columns associated with ss:
        standard_score: [49, 150]
        scaled_score:   [1,20]
        t_score:        [19, 80]
    """

    if ~np.isnan(percentile):
        standard_scores = DICT_PSYCHOMETRIC['standard_score'][percentile]
        scaled_scores = DICT_PSYCHOMETRIC['scaled_score'][percentile]
        t_scores = DICT_PSYCHOMETRIC['t_score'][percentile]
        
        if (value in standard_scores or
            (percentile == 1 and (value < 67 and value >= 32))):
        
            variable_values = [
                (rc_variables[1], value),
                (rc_variables[2], None),
                (rc_variables[3], None),
            ]
        elif (value in scaled_scores or 
              close_any(value, scaled_scores, 3) or
              (percentile == 1 and (value < 4 and value > 0))):
        
            variable_values = [
                (rc_variables[1], None),
                (rc_variables[2], value),
                (rc_variables[3], None),
            ]
        elif (value in t_scores or 
              close_any(value, t_scores, 3) or
              (percentile == 1 and (value < 28 and value > 0))):
        
            variable_values = [
                (rc_variables[1], None),
                (rc_variables[2], None),
                (rc_variables[3], value),
            ]
        else:
            raise PedsParserError()
    elif np.isnan(percentile):
        if value >= 80:
            variable_values = [
                (rc_variables[1], value),
                (rc_variables[2], None),
                (rc_variables[3], None),
            ]
        elif value >= 0 and value < 21:
            variable_values = [
                (rc_variables[1], None),
                (rc_variables[2], value),
                (rc_variables[3], None),
            ]
        elif value > 21 and value < 50:
            variable_values = [
                (rc_variables[1], None),
                (rc_variables[2], None),
                (rc_variables[3], value),
            ]
        else:
            raise PedsParserError()

    return variable_values
    

def peds_get_ss_variable(cell, rc_variables, percentile):
    """
    Returns the redcap variable and postprocessed value for 'ss' column
    
    Parameters
    ----------
    
    cell : Cell
        the cell from an openpyxl sheet
    rc_variables : list of str
        the list of redcap variables for the row of cell
    percentile : float
        the percentile value for the current row.
    
    Returns
    -------
    
    rc_variable : str
        the redcap variable name
    postprocessed_value : str or float
        the postprocessed cell value
    
    Description
    -----------
    
    Here are the lut columns associated with ss:
        standard_score
            between [49, 150]
        scaled_score
            between [1,20]
        t_score
            between [19, 80]
            the cell value may start with "T" or "T "
    """
    value = cell.value
    
    if value is not None and value not in neuroscore_parser.NAN_VALUES: 
        print(cell.row, cell.column_letter, value, cell.data_type)

        if cell.data_type == 'n':
            if cell.number_format == '"T"\\ 0;"T"\\ \\-0;"T"\\ 0;"T"\\ @':
                variable_values = [
                    (rc_variables[1], None),
                    (rc_variables[2], None),
                    (rc_variables[3], value),
                ]
            else:    
                variable_values = peds_determine_variable_value(value, rc_variables, percentile)
    
        elif cell.data_type == 's':
            if cell.number_format == '"T"\\ 0;"T"\\ \\-0;"T"\\ 0;"T"\\ @':
                postprocessed_value = float(re.sub('[<>]?[ ]*', '', value.strip()))

                variable_values = [
                    (rc_variables[1], None),
                    (rc_variables[2], None),
                    (rc_variables[3], postprocessed_value),
                ]
            elif re.fullmatch('T[ ]*[>]?\d+', value):
                postprocessed_value = float(re.sub('T[ ]*[>]?', '', value))
    
                variable_values = [
                    (rc_variables[1], None),
                    (rc_variables[2], None),
                    (rc_variables[3], postprocessed_value),
                ]

            elif re.fullmatch('[<>][ ]?\d+', value):
                value = float(re.sub('[<>][ ]?', '', value))
                variable_values = peds_determine_variable_value(value, rc_variables, percentile)

            else:
                variable_values = [(None, value)]
    
        else:
            variable_values = [
                (rc_variables[1], None),
                (rc_variables[2], None),
                (rc_variables[3], None),
            ]
    
    else:
        variable_values = [(None, value)]
    
    return variable_values

class neuroscore_parser():

    NAN_VALUES = {
        '#N/A',
        '#REF!',
        '#VALUE!',
        'RAW',
        'Raw',
        'Equivalent',
        'Form',
        'Notes',
        'Score',
        '',
        ' --',
        '%tile',
        '9-Min',
        'BLUE',
        'Can.',
        'FM-1',
        'L',
        'LIM.',
        None,
        'Norm',
        'Performance Indicator',
        'R',
        'SS',
        'STD',
        '[ERR]',
        '[FORM]',
        '[NORM]',
        '[SPAN]',
        '[TIME]',
        'raw',
        'val',
        '<Form>',
        '<Item Set>',
        '<Hand>',
        'Choose One',
    }

    def __init__(
        self, 
        wb_fname,
        sheet_name='Template', 
        form_info=None, 
        verbose=True
    ):
        """
        Initializes neuroscore_parser.

        Parameters
        ----------

        wb_fname: str
            path to excel workbook
        sheet_name: str
            the excel workbook sheet to parse
        form_info: (int, str, str)
            This variable holds information in order to replace tests matched with forms
            with a single test name in the parsed lines.
            First index is the form colun number (1-based).
            Second index is the form replacement file. Create this file to repalce test
            names with specific forms with a new single test name.
            The file format is expected to be a tsv file with these columns:
                test
                form
                new_test_name
            Third index is the valid form file. It is a workbook listing the valid forms
            to search under the form_name column.
        verbose : str
            be verboose about doing things

        Attributes
        ----------
            fname : str
                path to exel workbook
            wb : Workbook
                Workbook object for fname
            sh : Sheet
                Sheet object for "Template" in wb
            first_data_row : int
                first line in sh containing data (1-based)
            lines : list of line
                unique data entry lines in sh
            unhidden_lines : list of line
                line objects in lines that are unhidden in sh
        """

        self.fname = wb_fname
        self.wb = openpyxl.load_workbook(self.fname, data_only=True)
        self.sh = self.wb[sheet_name]

        self.first_data_row, self.first_data_col = self.find_first_data()
        if self.first_data_row > self.sh.max_row:
            raise Exception('first_data_row > sh.max_row')

        rd = self.sh.row_dimensions
        self.lines = self.parse_lines(form_info, verbose)
        self.unhidden_lines = [x for x in self.lines
                               if not rd[x.row].hidden]

    def find_first_data(self):
        """Returns the row, column for the first data entry"""

        for row, row_cells in enumerate(self.sh.iter_rows(), start=1):
            for col, cell in enumerate(row_cells):
                if cell.data_type == 's' and cell.value == 'Raw':
                    first_row = row + 1
                    first_col = col - 1
                    break
            else:
                continue
            break

        while self.sh[first_row][first_col].data_type == 'n':
            first_row = first_row + 1

        return (first_row, first_col)

    def _get_identifiers(self, verbose=True):
        """Returns unique data entry line identifiers in wb"""
        col = self.first_data_col

        output = defaultdict(list)

        # process first line here
        current_test = (
            self.sh[self.first_data_row][col].value, self.first_data_row
        )

        unique_stack = [self.sh[self.first_data_row][col].value]
        p_indent = self.sh[self.first_data_row][col].alignment.indent
        indent_mapper = {p_indent: 0}
        p_indent_key = p_indent

        output[current_test].append(
            (" | ".join(unique_stack), self.first_data_row)
        )

        start_row = self.first_data_row + 1
        for current_line, row in enumerate(
            self.sh.iter_rows(min_row=start_row), start=start_row
        ):
            if (row and
                row[col].value and
                row[col].value.strip() and
                not row[col].value.strip().startswith('*')):

                c_text = row[col].value

                c_indent = row[col].alignment.indent
                if c_text.startswith(' '):
                    c_indent = c_indent + 1
                if c_indent == 0 and not row[col].font.b:
                    c_indent = 1

                if c_indent not in indent_mapper:
                    indent_mapper[c_indent] = indent_mapper[p_indent_key] + 1
                    p_indent_key = c_indent
                c_indent = indent_mapper[c_indent]

                if verbose:
                    print(f"Parsing line : {current_line} : {c_text} : {c_indent}")

                c_text = c_text.strip()

                if c_indent == p_indent:
                    unique_stack.pop()
                    unique_stack.append(c_text)

                    if c_indent == 0:
                        current_test = (c_text, current_line)
                        
                elif c_indent > p_indent:
                    unique_stack.append(c_text)
                else:
                    if c_indent == 0:
                        unique_stack.clear()

                        current_test = (c_text, current_line)

                        indent_mapper = {c_indent: 0}
                        p_indent_key = c_indent
                    else:
                        unique_stack.pop()
                        unique_stack.pop()

                    unique_stack.append(c_text)

                output[current_test].append(
                    (" | ".join(unique_stack), current_line)
                )

                p_indent = c_indent

        return output

    def _replace_testforms(self, unique_identifiers, form_info):
        """
        Finds grabs the forms for the tests in unique_identifiers.

        Parameters
        ----------

        unique_identifiers: dict ((test_identifier, line_no) : [unique_identifiers])
            The unique_identifiers mapped to their test name.
        valid_forms: list
            The list of valid forms to identifiy for tests.
        form_column: str
            Search this excel column for forms.

        Returns
        -------

        test_forms: dict ((test_identifier, line_no) : str)
            The form associated with the test. The empty string indicates not form
            found.
        """

        replaced_uids = dict()
        form_column = form_info[0]
        form_lut = Path(form_info[1]).resolve()
        file_valid_forms = Path(form_info[2]).resolve()
        replacements = (read_workbook(form_lut)
            .loc[lambda df_: df_.replacement.notnull()]
        )
        valid_forms = set(read_workbook(file_valid_forms).form_name)
        replacements.set_index(['test', 'form'], inplace=True)
        for test, lines in unique_identifiers.items():
            forms = []
            min_row = lines[0][1]
            max_row = lines[-1][1] 
            for row in self.sh.iter_rows(
                min_row, 
                max_row, 
                form_column, 
                form_column, 
                True
            ):
                if row[0] in valid_forms:
                    forms.append(row[0])

            check = (test[0], ' | '.join(forms))
            if check in replacements.index:
                replacement = replacements.loc[check, 'replacement']
                replaced_uids[(replacement, test[1])] = [
                    (x[0].replace(test[0], replacement, 1), x[1])
                    for x in lines
                ]
            else:
                replaced_uids[test] = lines

        return replaced_uids

    def parse_lines(self, form_info, verbose=True):

        output = []
        unique_identifiers = self._get_identifiers(verbose)
        if form_info:
            unique_identifiers = self._replace_testforms(
                unique_identifiers, 
                form_info
            )

        test_counter = defaultdict(int)
        for test_info, uids in unique_identifiers.items():
            test_counter[test_info[0]] += 1
            for uid in uids:
                output.append(line(uid[0], test_counter[test_info[0]], uid[1]))

        return output

    def get_test_forms(self, form_column, verbose=True):
        """Returns all tests and their forms for original identifiers"""

        results = {
            'line_no': [],
            'test': [],
            'form': [],
        }
        unique_identifiers = self._get_identifiers(verbose)
        for test, lines in unique_identifiers.items():
            forms = []
            min_row = lines[0][1]
            max_row = lines[-1][1] 
            for row in self.sh.iter_rows(
                min_row, 
                max_row, 
                form_column, 
                form_column, 
                True
            ):
                if row[0] and row[0] not in self.NAN_VALUES:
                    forms.append(row[0])

            if forms:
                results['line_no'].append(test[1])
                results['test'].append(test[0])
                results['form'].append(' | '.join(forms))

        return pd.DataFrame(results)

    def find_new_lines(self, lut):
        """Return new lines in sh not found in lut"""

        new_lines = [line for identifier, test_no, row in self.lines
                     if identifier not in lut.lut]

        return new_lines

    def find_administered_tests(self):
        """Returns the administered test found in sh"""

        return [identifier for identifier, _, _ in self.unhidden_lines
                if identifier.find(' | ') == -1]

    def parse_data(self, lut, tp):
        """
        Parse the data in sh using lut for timepoint tp

        Parameters
        ----------

        lut : lut
            The lookup table for parsing.
        tp : int
            The timepoint number
            Used only for epilepsy, dementia, or aphasia depts

        Returns
        -------

        results : dict (variable : [value])
            parsed neuroscore values according to lut
        new_lines : dict
            Each entry contains a new identifier/test_no information
            test       -> list
            test_no    -> list
            identifier -> list
            row        -> list
        missing_lines : dict
            Contains neuroscore variables where no redcap variable is assigned,
            but an identifier is present in the lut
            test       -> list
            test_no    -> list
            identifier -> list
            row        -> list
            col        -> list
            name       -> list of neuroscore column names
            value      -> neuroscore value
        """

        data_cols = ['raw', 'ss', 'percentile', 'notes']
        tp_offset = 4 * (tp - 1)
        col_offset = 2 + tp_offset
            
        results = {}
        new_lines = {
            'test': [],
            'test_no': [],
            'identifier': [],
            'row': [],
        }
        missing_lines = {
            'test': [],
            'test_no': [],
            'identifier': [],
            'row': [],
            'col': [],
            'name': [],
            'value': [],
        }

        for identifier, test_no, row in self.unhidden_lines:
            key = (identifier, test_no)
            if key in lut.lut:
                rc_variables = lut.lut[key]

                for n, variable in enumerate(rc_variables):
                    try:
                        value = self.sh[row][n + col_offset].value
                        if value in self.NAN_VALUES:
                            value = np.nan
                    except:
                        value = np.nan

                    if variable:
                        results[variable] = [value]
                    elif not pd.isna(value):
                        missing_lines['test'].append(identifier.split(' | ')[0])
                        missing_lines['test_no'].append(test_no)
                        missing_lines['identifier'].append(identifier)
                        missing_lines['row'].append(row)
                        missing_lines['col'].append(n + col_offset)
                        missing_lines['name'].append(data_cols[n])
                        missing_lines['value'].append(value)
                        
            else:
                new_lines['test'].append(identifier.split(' | ')[0])
                new_lines['test_no'].append(test_no)
                new_lines['identifier'].append(identifier)
                new_lines['row'].append(row)
                    
        return results, new_lines, missing_lines

    def parse_header(self, tp, study):
        """Parse header information for timepoint tp and study study"""

        results = {}

        if self.dept in {'peds'}:
            results['Provider'] = self.sh['C4'].value
            results['Psychometrist'] = self.sh['C5'].value

            results['Sex'] = self.sh['G4'].value
            results['DOE'] = self.sh['G5'].value
            results['DOB'] = self.sh['G6'].value
            results['Yrs'] = self.sh['G7'].value
            results['Mo']  = self.sh['G8'].value
            results['D']   = self.sh['G9'].value
            results['Handedness'] =  self.sh['G10'].value

        elif self.dept in {'epilepsy', 'dementia', 'aphasia'}:
            date_col = 4 + (tp - 1) * 4
            age_col = 2 + (tp -1) * 4

            results['testdat'] = self.sh[9][date_col].value.strftime('%Y-%m-%d')
            results['age'] = [int(self.sh[11][age_col].value
                                  .split(',')[0]
                                  .split(': ')[1])]
        else:
            raise Exception(f'Unkown dept: {dept}')

        return results

class mci_parser(neuroscore_parser):

    def parse_date(self, tp):
        """Retrieve the date for a timepoint"""

        for row in range(1, self.sh.max_row + 1):
            for col, cell in enumerate(self.sh[row]):
                if (cell.data_type == 's' and 
                    (cell.value.startswith('DOS A:') or 
                     cell.value.startswith('EXAM A:') or
                     cell.value.startswith('Exam A:'))):

                    first_row = row
                    first_col = col
                    break
            else:
                continue
            break

        if cell.value.startswith('DOS A:'):
            row = first_row + tp - 1
            txt = self.sh[row][first_col].value.split(':')[1].strip()

            exam_date = datetime.strptime(txt, '%B %d, %Y').strftime('%Y-%m-%d')
        elif cell.value.startswith('EXAM A:'):
            col = (first_col + 2) + ((tp - 1) * 4)
            exam_date = self.sh[first_row][col].value.strftime('%Y-%m-%d')
        elif cell.value.startswith('Exam A'):
            col = first_col + ((tp - 1) * 4)
            txt = self.sh[first_row][col].value.strip().split()
            exam_date = datetime.strptime(txt[-1], '%d-%b-%Y').strftime('%Y-%m-%d')
        else:
            exam_date = None

        return exam_date
            
class peds_parser(neuroscore_parser):

    def parse_data(self, lut):
        """
        Parse the data in sh using lut

        Parameters
        ----------

        lut : lut
            The lookup table for parsing.

        Returns
        -------

        results : dict (variable : [value])
            parsed neuroscore values according to lut
        new_lines : dict
            Each entry contains a new identifier/test_no information
            identifier -> list
            test_no    -> list
            row        -> list
        missing_lines : dict
            Contains neuroscore variables where no redcap variable is assigned,
            but an identifier is present in the lut
            identifier -> list
            test_no    -> list
            row        -> list
            col        -> list
            name       -> list of neuroscore column names
            value      -> neuroscore value
        """

        data_cols = ['raw', 'ss', '%tile', 'equivalent', 'form', 'notes']
        get_variables = [
            (0, self._get_raw_variable),
            (1, self._get_ss_variable),
            (2, self._get_percentile_variable),
            (3, self._get_equivalent_variable),
            (4, self._get_form_variable),
            (5, self._get_notes_variable),
        ]
            
        results = self.parse_header()
        debug_results = {
            'identifier': [],
            'variable': [],
            'value': [],
        }
        new_lines = {
            'test': [],
            'test_no': [],
            'identifier': [],
            'row': [],
        }
        missing_lines = {
            'test': [],
            'test_no': [],
            'identifier': [],
            'row': [],
            'col': [],
            'col_letter': [],
            'name': [],
            'value': [],
        }

        for identifier, test_no, row in self.unhidden_lines:
            key = (identifier, test_no)
            if key in lut.lut:
                rc_variables = lut.lut[key]

                for n, get_variable in get_variables:

                    cell = self.sh[row][n + self.first_data_col + 1]
                    variable_values = get_variable(cell, rc_variables)

                    for variable, value in variable_values:
                        is_nan_value = value in self.NAN_VALUES

                        if variable and not is_nan_value:
                            results[variable] = [value]

                            debug_results['identifier'].append(identifier)
                            debug_results['variable'].append(variable)
                            debug_results['value'].append(value)
                        elif not pd.isna(value) and not is_nan_value:
                            missing_lines['test'].append(identifier.split(' | ')[0])
                            missing_lines['test_no'].append(test_no)
                            missing_lines['identifier'].append(identifier)
                            missing_lines['row'].append(row)
                            missing_lines['col'].append(cell.column)
                            missing_lines['col_letter'].append(cell.column_letter)
                            missing_lines['name'].append(data_cols[n])
                            missing_lines['value'].append(value)

            else:
                new_lines['test'].append(identifier.split(' | ')[0])
                new_lines['test_no'].append(test_no)
                new_lines['identifier'].append(identifier)
                new_lines['row'].append(row)

        return results, debug_results, new_lines, missing_lines

    def _get_raw_variable(self, cell, rc_variables):
        """Returns the redcap variable and postprocessed value for raw column"""

        return [(rc_variables[0], cell.value)]

    def _get_ss_variable(self, cell, rc_variables):
        """
        Returns the redcap variable and postprocessed value for 'ss' column

        Parameters
        ----------

        cell : Cell
            the cell from an openpyxl sheet
        rc_variables : list of str
            the list of redcap variables for the row of cell


        Returns
        -------

        rc_variable : str
            the redcap variable name
        postprocessed_value : str or float
            the postprocessed cell value

        Description
        -----------

        Here are the lut columns associated with ss:
            standard_score
                between (60, 120]
            scaled_score
                between [1,35]
            t_score
                between [40, 60]
                the cell value may start with "T" or "T "
        """

        return [(rc_variables[1], cell.value)]

    def _get_percentile_variable(self, cell, rc_variables):
        """Returns the redcap variable and postprocessed value for 'percentile' colum"""

        return [(rc_variables[2], cell.value)]

    def _get_equivalent_variable(self, cell, rc_variables):
        """
        Returns the redcap variables and postprocessed values for 'equivalent' column.

        Parameters
        ----------

        cell : Cell
            the cell from an openpyxl sheet
        rc_variables : list of str
            the list of redcap variables for the row of cell


        Returns
        -------

        rc_variable : str
            the redcap variable name
        postprocessed_value : str or float
            the postprocessed cell value

        Description
        -----------
        The eqiuvalent column is the only column that can return multiple values.
        Here are the columns associated with 'equivalent':
            sign
                <, > will be listed or two values with a dash will be listed
            age_equivalent
                either enter the singal valeu listed or the first value ifa  range
            high_equivalent
                either leave blank if a single value listed or enter the send value of range
        """

        value = cell.value

        if value is not None:
            if cell.data_type == 's':
                if re.match('[<>]', value):
                    variable_values = [
                        (rc_variables[3], value),
                        (rc_variables[4], None),
                        (rc_variables[5], None)
                    ]
                elif re.fullmatch('\d+-\d+', value):
                    variable_values = [
                        (rc_variables[3], value),
                        (rc_variables[4], None),
                        (rc_variables[5], None),
                    ]
                elif re.fullmatch('\d+:\d+', value):
                    m = re.fullmatch('(\d+):(\d+)', value)
                    n1 = float(m.group(1))
                    n2 = float(m.group(2))
                    variable_values = [
                            (rc_variables[3], None),
                            (rc_variables[4], n1 * 12 + n2),
                            (rc_variables[5], None),
                    ]
                elif re.fullmatch('\d+:\d+-\d+:\d+', value):
                    m = re.fullmatch('(\d+):(\d+)-(\d+):(\d+)', value)
                    n1 = float(m.group(1))
                    n2 = float(m.group(2))
                    n3 = float(m.group(3))
                    n4 = float(m.group(4))
                    variable_values = [
                            (rc_variables[3], None),
                            (rc_variables[4], n1 * 12 + n2),
                            (rc_variables[5], n3 * 12 + n4),
                    ]
                else:
                    variable_values = [(None, value)]

            elif cell.data_type == 'n':
                variable_values = [
                    (rc_variables[3], None),
                    (rc_variables[4], float(value)),
                    (rc_variables[5], None),
                ]

            else:
                variable_values = [(None, value)]

        else:
            variable_values = [
                (rc_variables[3], None),
                (rc_variables[4], None),
                (rc_variables[5], None),
            ]

        return variable_values

    def _get_form_variable(self, cell, rc_variables):
        """Returns the redcap variable and postprocessed value for 'form' column"""

        return [(rc_variables[7], cell.value)]

    def _get_notes_variable(self, cell, rc_variables):
        """Returns the redcap variable and postprocesse value for the 'notes' column"""

        return [(rc_variables[8], cell.value)]

    def parse_header(self):
        """Returns the header information"""

        header = {
            'mrn': [np.nan],
            'doe': [np.nan],
            'dob': [np.nan],
            'years': [np.nan],
            'months': [np.nan],
            'days': [np.nan],
            'gender': [np.nan],
            'handedness': [np.nan],
        }

        key_mapper = {
            'yrs': 'years',
            'mo': 'months',
            'd': 'days',
        }

        for row in range(1, self.first_data_row):
            col = 0
            ncols = len(self.sh[row])
            while col < ncols:
                if self.sh[row][col].data_type == 's' and self.sh[row][col].value.endswith(':'):
                    key = self.sh[row][col].value[:-1].lower().strip()
                    if key in key_mapper:
                        key = key_mapper[key]
                    col += 1
                    while col < ncols and not self.sh[row][col].value:
                        col += 1
                    if col < ncols and self.sh[row][col].value:
                        header[key] = [self.sh[row][col].value]

                col += 1

        return header

class neonatology_parser(neuroscore_parser):

    def parse_data(self, lut):
        """
        Parse the data in sh using lut

        Parameters
        ----------

        lut : lut
            The lookup table for parsing.

        Returns
        -------

        results : dict (variable : [value])
            parsed neuroscore values according to lut
        new_lines : dict
            Each entry contains a new identifier/test_no information
            identifier -> list
            test_no    -> list
            row        -> list
        missing_lines : dict
            Contains neuroscore variables where no redcap variable is assigned,
            but an identifier is present in the lut
            identifier -> list
            test_no    -> list
            row        -> list
            col        -> list
            name       -> list of neuroscore column names
            value      -> neuroscore value
        """

        data_cols = ['raw', 'ss', 'percentile', 'equivalent', 'form', 'notes', 'gsv']
            
        results = self.parse_header()
        debug_results = {
            'identifier': [],
            'variable': [],
            'value': [],
        }
        new_lines = {
            'test': [],
            'test_no': [],
            'identifier': [],
            'row': [],
        }
        missing_lines = {
            'test': [],
            'test_no': [],
            'identifier': [],
            'row': [],
            'col': [],
            'col_letter': [],
            'name': [],
            'value': [],
        }

        for identifier, test_no, row in self.unhidden_lines:
            key = (identifier, test_no)
            if key in lut.lut:
                rc_variables = lut.lut[key]

                for n, variable in enumerate(rc_variables):
                    cell = self.sh[row][n + self.first_data_col + 1]
                    value = cell.value
                    if value in self.NAN_VALUES:
                        value = np.nan

                    if variable:
                        results[variable] = [value]

                        debug_results['identifier'].append(identifier)
                        debug_results['variable'].append(variable)
                        debug_results['value'].append(value)
                    elif not pd.isna(value):
                        missing_lines['test'].append(identifier.split(' | ')[0])
                        missing_lines['test_no'].append(test_no)
                        missing_lines['identifier'].append(identifier)
                        missing_lines['row'].append(row)
                        missing_lines['col'].append(cell.column)
                        missing_lines['col_letter'].append(cell.column_letter)
                        missing_lines['name'].append(data_cols[n])
                        missing_lines['value'].append(value)
                        
            else:
                new_lines['test'].append(identifier.split(' | ')[0])
                new_lines['test_no'].append(test_no)
                new_lines['identifier'].append(identifier)
                new_lines['row'].append(row)
                    
        return results, debug_results, new_lines, missing_lines

    def parse_header(self):
        """Returns the header information"""

        header = {
            'mrn': [np.nan],
            'doe': [np.nan],
            'dob': [np.nan],
            'years': [np.nan],
            'months': [np.nan],
            'days': [np.nan],
            'sex': [np.nan],
            'handedness': [np.nan],
            'dominant_language': [np.nan],
            'interpreter_used': [np.nan],
            'adjusted_age_years': [np.nan],
            'adjusted_age_months': [np.nan],
            'adjusted_age_days': [np.nan],
        }

        key_mapper = {
            'yrs': 'years',
            'mo': 'months',
            'd': 'days',
            'dominant language': 'dominant_language',
            'interpreter used': 'interpreter_used',
            'adjusted_age_yrs': 'adjusted_age_years',
            'adjusted_age_mo': 'adjusted_age_months',
            'adjusted_age_d': 'adjusted_age_days',
        }

        adjusted_age_col = 0
        adjusted_age_row = 0
        iter_rows = self.sh.iter_rows(
            max_row=self.first_data_row, 
            values_only=True
        )
        for i, row in enumerate(iter_rows, start=1):
            for j, col in enumerate(row, start=1):
                if col == 'Adjusted Age':
                    adjusted_age_row = i
                    adjusted_age_col = j
                    break
            else:
                continue
            break

        max_col = (adjusted_age_col - 1) if adjusted_age_col else None
        for row in self.sh.iter_rows(max_row=self.first_data_row, max_col=max_col):
            for i, col in enumerate(row):
                if col.data_type == 's' and col.value.endswith(':'):
                    key = col.value[:-1].lower().strip()
                    if key in key_mapper:
                        key = key_mapper[key]
                    for cell_val in row[(i + 1):]:
                        if cell_val.value is not None:
                            if cell_val.value in self.NAN_VALUES:
                                header[key] = [np.nan]
                            else:
                                header[key] = [cell_val.value]
                            break

        if adjusted_age_col > 0:
            iter_rows = self.sh.iter_rows(
                min_row=adjusted_age_row,
                max_row=self.first_data_row,
                min_col=adjusted_age_col,
            )
            for row in iter_rows:
                field = row[0]
                if field.data_type == 's' and field.value.endswith(':'):
                    key = 'adjusted_age_' + field.value[:-1].lower().strip()
                    if key in key_mapper:
                        key = key_mapper[key]
                    for cell_val in row[1:]:
                        if cell_val.value is not None:
                            if cell_val.value in self.NAN_VALUES:
                                header[key] = [np.nan]
                            else:
                                header[key] = [cell_val.value]

        return header
