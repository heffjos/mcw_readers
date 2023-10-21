import openpyxl

import pandas as pd

def initialize_lut(excel, dept):
    """
    Initialize a blank lookup table for neuroscore file for a specific department.

    **Parameters**

        excel
            Neuroscore file name.
        dept
            A str indicating what the department for the excel file.
            Here are the available choices:
                peds
                epilepsy
                dementia
                aphasia

    **Outputs**

        output
            A dataframe with the following columns:
                test - the test name the row belongs
                test_no - the current test number
                identifier - the unique identifier for the row
    """
    output = {'test': [], 'test_no': [], 'identifier': []}
    test_counter = {}

    if dept in {'peds'}:
        TEST_COL = 2
        RAW_COL = 3
    elif dept in {'epilepsy', 'dementia', 'aphasia'}:
        TEST_COL = 1
        RAW_COL = 2
    else:
        raise Exception(f'Unkown dept: {dept}')

    wb = openpyxl.load_workbook(excel, data_only=True)
    sh = wb['Template']

    begin_line = 1
    while ((not sh[begin_line] or 
        not sh[begin_line][RAW_COL].value or
        sh[begin_line][RAW_COL].value.strip() != "Raw") and 
        begin_line < sh.max_row):

        begin_line = begin_line + 1

    while ((not sh[begin_line] or 
        not sh[begin_line][TEST_COL].value) and
        begin_line < sh.max_row):

        begin_line = begin_line + 1

    print(f'begin_line : {begin_line}')
            
    if begin_line < sh.max_row:

        current_test = sh[begin_line][TEST_COL].value
        if current_test in test_counter:
            test_counter[current_test] += 1
        else:
            test_counter[current_test] = 1

        unique_stack = [sh[begin_line][TEST_COL].value]
        p_indent = sh[begin_line][TEST_COL].alignment.indent
        indent_mapper = {p_indent: 0}
        p_indent_key = p_indent

        output['identifier'].append(' | '.join(unique_stack))
        output['test'].append(current_test)
        output['test_no'].append(test_counter[current_test])

        for current_line in range(begin_line + 1, sh.max_row + 1):

            if (sh[current_line] and 
                sh[current_line][TEST_COL].value and 
                sh[current_line][TEST_COL].value.strip() and
                not sh[current_line][TEST_COL].value.strip().startswith('*')):

                c_text = sh[current_line][TEST_COL].value

                c_indent = sh[current_line][TEST_COL].alignment.indent
                if c_text.startswith(' '):
                    c_indent = c_indent + 1

                if c_indent not in indent_mapper:
                    indent_mapper[c_indent] = indent_mapper[p_indent_key] + 1
                    p_indent_key = c_indent
                c_indent = indent_mapper[c_indent]

                print(f"Parsing line : {current_line} : {c_text} : {c_indent}")

                c_text = c_text.strip()

                if c_indent == p_indent:
                    unique_stack.pop()
                    unique_stack.append(c_text)

                    if c_indent == 0:
                        current_test = c_text
                        if current_test in test_counter:
                            test_counter[current_test] += 1
                        else:
                            test_counter[current_test] = 1
                        
                elif c_indent > p_indent:
                    unique_stack.append(c_text)
                else:
                    if c_indent == 0:
                        unique_stack.clear()

                        current_test = c_text
                        if current_test in test_counter:
                            test_counter[current_test] += 1
                        else:
                            test_counter[current_test] = 1

                        indent_mapper = {c_indent: 0}
                        p_indent_key = c_indent
                    else:
                        unique_stack.pop()
                        unique_stack.pop()

                    unique_stack.append(c_text)

                output['identifier'].append(' | '.join(unique_stack))
                output['test'].append(current_test)
                output['test_no'].append(test_counter[current_test])

                p_indent = c_indent

    return output

class lut():

    def __init__(self, dept, excel, sheet_name=0):

        self.excel = excel

        if excel.endswith('csv'):
            self.df = pd.read_csv(excel)
        else:
            self.df = pd.read_excel(self.excel, sheet_name = sheet_name)

        self.split_identifiers = [x.split(' | ') 
                                  for x in self.df['identifier']]

        if dept not in {'peds', 'epilepsy', 'dementia', 'aphasia', 'neonatology'}:
            raise Exception(f'Unknown dept: {dept}')
        self.dept = dept

        self.lut = self.convert_df_to_lut()

    def convert_df_to_lut(self):
        """Converts a df lut to a dict lut"""

        if self.dept in {'peds'}:
            data_cols = [
                'raw', 
                'ss',
                'percentile', 
                'sign',
                'age_equivalent',
                'high_equivalent',
                'developmental_quotient',
                'form', 
                'notes',
            ]
        elif self.dept in {'neonatology'}:
            data_cols = [
                'raw',
                'ss',
                'percentile',
                'equivalent',
                'form',
                'notes',
                'gsv'
            ]
        elif self.dept in {'epilepsy', 'dementia', 'aphasia'}:
            data_cols = ['raw', 'ss', 'percentile', 'notes']
        else:
            raise Exception(f'Unkown dept: {dept}')

        values = self.df.fillna({x:'' for x in data_cols})[data_cols].values.tolist()
        results = {(identifier, test_no): values 
                   for identifier, test_no, values 
                   in zip(self.df['identifier'], 
                          self.df['test_no'], 
                          values)}

        return results

    def get_headers_at_indent_level(self, level):
        """Returns all headers at indent level `level`"""

        return ['' if len(x) < (level + 1) else x[level] 
                for x in self.split_identifiers]

